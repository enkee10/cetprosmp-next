import json
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `python -m pip install openpyxl`."
    ) from exc


SHEET_NAME = "Datos extraidos"


def cell_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def color_key(cell):
    fg = cell.fill.fgColor
    if fg.type == "rgb":
        return fg.rgb
    if fg.type == "theme":
        return f"theme:{fg.theme}:tint:{fg.tint}"
    return fg.type or "none"


def row_color(row):
    colors = Counter(color_key(cell) for cell in row)
    if colors.get("FFFFFF00"):
        return "yellow"
    if any(key.startswith("theme:7:") for key in colors):
        return "blue"
    if colors.get("FF0F766E"):
        return "header"
    return "white"


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python scripts/parse-academic-excel.py <xlsx-path>")

    workbook_path = Path(sys.argv[1])
    if not workbook_path.exists():
        raise SystemExit(f"Excel file not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")

    worksheet = workbook[SHEET_NAME]
    raw_headers = [str(cell.value or "").strip() for cell in worksheet[1]]
    headers = [
        header if header else f"col_{index + 1}"
        for index, header in enumerate(raw_headers)
    ]

    rows = []
    color_counts = Counter()
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        values = {
            headers[index]: cell_value(cell.value)
            for index, cell in enumerate(row)
            if index < len(headers)
        }
        if not any(value not in (None, "") for value in values.values()):
            continue

        color = row_color(row)
        color_counts[color] += 1
        rows.append(
            {
                "rowNumber": row_number,
                "color": color,
                "values": values,
            }
        )

    payload = json.dumps(
        {
            "path": str(workbook_path),
            "sheet": SHEET_NAME,
            "headers": headers,
            "colorCounts": dict(color_counts),
            "rows": rows,
        },
        ensure_ascii=False,
    )
    sys.stdout.buffer.write(payload.encode("utf-8"))


if __name__ == "__main__":
    main()
